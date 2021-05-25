from main import HOST
import openpyxl
from bs4 import BeautifulSoup as bs
import requests
from PIL import Image
from io import BytesIO


wb = openpyxl.Workbook()
wb.create_sheet(title='List1', index=0)
sheet = wb['List1']
sheet.append(['Наименование', 'Артикул', 'Цена(руб.)', 'Описание'])


def in_catalog_page():
    ff = 'https://www.stout.ru/catalog/aksessuary-i-komplektuyushchie'
    #ff = 'https://www.stout.ru/catalog/vodonagrevateli-0'
    page_item = ''
    response_page = requests.get(ff)
    page = bs(response_page.text, 'lxml')
    bb = page.find_all('div', class_='field field-name-node-link field-type-ds field-label-hidden')
    for p_item in bb:
        pg = p_item.find('a').get('href')
        page_item += f"{HOST}{pg},"
    len_page_item = len(page_item.split(','))
    for i in range(len_page_item):
        print(page_item.split(',')[i])
        if page_item.split(',')[i] != "":
            page_item_path = requests.get(page_item.split(',')[i])
            page_item_cat = bs(page_item_path.text, 'lxml')
            item_title = page_item_cat.find('h1', class_='page-header')
            item_artcles = page_item_cat.find('div', class_='field-item even', itemprop='sku')
            item_price_up_level = page_item_cat.find('div', itemprop='offers')
            item_content = page_item_cat.find_all(class_='views-field views-field-field-char')
            item_img = page_item_cat.find('img', {"class":"image field-slideshow-image field-slideshow-image-1 img-responsive"})
            print(item_img['src'])
            #img_to_e = openpyxl.drawing.image.Image(item_img_resize(item_img['src']))
            rec_img_to_ex(sheet, item_img)
            print("Название: {}".format(item_title.text))
            print("Артикул: {}".format(item_artcles.text))
            print('Цена: {}p'.format(item_price_up_level.text[:-1]))
            opis = []
            opis.append(item_title.text)
            opis.append(item_artcles.text)
            opis.append(item_price_up_level.text[:-1])
            for i in item_content:
                str_a = i.find_all('div', class_='field-label')
                str_b = i.find_all('div', class_='field-items')
                for s_a, s_b in zip(range(len(str_a)), range(len(str_b))):
                    print(str_a[s_a].text, str_b[s_b].text)
                    opis.append(str_a[s_a].text.replace('\xa0', ' ') + str_b[s_b].text.replace('\xa0', ' ') + " ")

            #sheet.append(opis)
            #sheet.add_image(img_to_e)


def item_img_resize(url, size=(100, 100)):
    r = requests.get(url, stream=True)
    r.raw.decode_content = True
    img = Image.open(r.raw)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format='png')
    temp.seek(0)
    return Image.open(temp)

def rec_img_to_ex(sheet, img_url, size=(100, 100)):
    img = openpyxl.drawing.image.Image(item_img_resize(img_url['src']))
    row_n = sheet.max_row + 1
    cell_addr = f'A{row_n}'
    sheet.add_image(img)
    sheet.row_dimensions[row_n].height = int(size[1]* .8)
    sheet.column_dimensions['A'].width = int(size[0] * .2)

print(in_catalog_page())
wb.save('book.xls')




