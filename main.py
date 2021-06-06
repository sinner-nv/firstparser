import openpyxl
from bs4 import BeautifulSoup as bs
import requests
from PIL import Image
from io import BytesIO


HOST = 'https://www.stout.ru'
PG_I = '?page='
response = requests.get(HOST)
soup = bs(response.text, 'lxml')
wb = openpyxl.Workbook()



def category():
    dontentr = ['Контакты',
            'О бренде',
            'Карта сайта',
            'Где купить',
            'Обратная связь',
            'Выполненные объекты',
            'Статьи']
    category = ''
    q = soup.find_all('li', class_='leaf')
    for i, cat in enumerate(q, start=1):
        if cat.text not in dontentr:
            category += f"{HOST}{cat.find('a').get('href')},"
    return category.split(',')

def in_catalog_page(ff, sheet):
    #ff = 'https://www.stout.ru/catalog/aksessuary-i-komplektuyushchie'
    #ff = 'https://www.stout.ru/catalog/vodonagrevateli-0'
    page_item = ''
    response_page = requests.get(ff)
    page = bs(response_page.text, 'lxml')
    bb = page.find_all('div', class_='field field-name-node-link field-type-ds field-label-hidden')
    for p_item in bb:
        pg = p_item.find('a').get('href')
        page_item += f"{HOST}{pg},"
    len_page_item = len(page_item.split(','))
    for i in range(len_page_item):
        print(page_item.split(',')[i])
        if page_item.split(',')[i] != "":
            page_item_path = requests.get(page_item.split(',')[i])
            page_item_cat = bs(page_item_path.text, 'lxml')
            item_title = page_item_cat.find('h1', class_='page-header')
            item_artcles = page_item_cat.find('div', class_='field-item even', itemprop='sku')
            item_price_up_level = page_item_cat.find('div', itemprop='offers')
            item_content = page_item_cat.find_all(class_='views-field views-field-field-char')
            item_img = page_item_cat.find('img',
                                          {"class": "image field-slideshow-image field-slideshow-image-1 img-responsive"})
            #print(item_img['src'])
            print("Название: {}".format(item_title.text))
            print("Артикул: {}".format(item_artcles.text))
            print('Цена: {}p'.format(item_price_up_level.text[:-1]))
            opis = []
            for i in item_content:
                str_a = i.find_all('div', class_='field-label')
                str_b = i.find_all('div', class_='field-items')
                for s_a, s_b in zip(range(len(str_a)), range(len(str_b))):
                    print(str_a[s_a].text, str_b[s_b].text)
                    opis.append(str_a[s_a].text.replace('\xa0', ' ') + str_b[s_b].text.replace('\xa0', ' ') + " ")

            rec_img_to_ex(sheet,
                          item_img,
                          item_title.text,
                          item_artcles.text,
                          item_price_up_level.text[:-1],
                          opis)

def item_img_resize(url, size=(100, 100)):
    r = requests.get(url, stream=True)
    r.raw.decode_content = True
    img = Image.open(r.raw)
    if size:
        img = img.resize(size)
    temp = BytesIO()
    img.save(temp, format='png')
    temp.seek(0)
    return Image.open(temp)

def rec_img_to_ex(sheet, img_url, item_title, item_artcles, item_price, opis, size=(100, 100)):
    img = openpyxl.drawing.image.Image(item_img_resize(img_url['src']))
    row_n = sheet.max_row + 1
    cell_addr = f'A{row_n}'
    img.anchor = cell_addr
    sheet.add_image(img)
    sheet[f"B{row_n}"] = item_title
    sheet[f"C{row_n}"] = item_artcles
    sheet[f"D{row_n}"] = item_price
    col = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V',
           'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL',
           'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA']
    for i in range(len(opis)):
        sheet[f"{col[i]}{row_n}"] = opis[i]
        sheet.column_dimensions[f"{col[i]}"].width = int(size[0] * .6)
    sheet.row_dimensions[row_n].height = int(size[1] * .8)
    sheet.column_dimensions['A'].width = int(size[0] * .2)
    sheet.column_dimensions['B'].width = int(size[0] * .6)
    sheet.column_dimensions['C'].width = int(size[0] * .3)
    sheet.column_dimensions['D'].width = int(size[0] * .2)

def bok():
    index_page = 0
    for i in category():
        if i !='':
            response_page = requests.get(i)
            page = bs(response_page.text, 'lxml')
            bb = page.find_all('h1', class_='page-header')
            for i in bb:
                print(category()[index_page])
                wb.create_sheet(title=i.text, index=index_page)
                sheet = wb[i.text]
                sheet.append(['Изображение', 'Наименование', 'Артикул', 'Цена(руб.)'])
                for number_page in range(0, 100):
                    pg_las = f"{category()[index_page]}{PG_I}{number_page}"
                    resp_page = requests.get(pg_las)
                    page_n = bs(resp_page.text, 'lxml')
                    pg_item = page_n.find('li', class_='next')
                    print("Страница: ", pg_las)
                    print("Номер страницы:", number_page)
                    if (pg_item is None):
                        continue
                    else:
                        in_catalog_page(pg_las, sheet)
                        index_page +=1

                # in_catalog_page(category()[index_page], sheet)
                # index_page +=1



bok()
wb.save('book.xls')